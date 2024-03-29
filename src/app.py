import time
import json
from pathlib import Path

from flask import Flask, request, redirect, url_for, render_template, send_file # python web framework
from tinydb import TinyDB, table # json database management
from openpyxl import Workbook # for Excel exportation


app = Flask(__name__) # flask app initialization

db = TinyDB("db.json", indent=4).table("evaluations") # database initialization for students and her marks 
db_criteria = TinyDB(Path(__file__).parent/"_db.json", indent=4).table("criteria")  # db for criteria

# api request status
status = {
    "success": {
        "status": 0,
        "object": None,
    },
    "error": {
        "status": 1,
        "error_message": ""
    }
}

# utilities functions

def evaluate(student, criteria, value) -> dict:

    evaluations = student.get("evaluations") # evaluations list 
    
    evaluation_object = None 
    
    # get criteria
    for i in range(len(evaluations)) :
        if evaluations[i]["criteria"] == criteria["name"]:
            evaluation_object = evaluations.pop(i)
            break
        
    if evaluation_object is not None:
        evaluation_object["values"].append(value)
    else:
        # add new criteria 
        evaluation_object = {
            "criteria": criteria["name"],
            "values": [value]
        }

    evaluations.append(evaluation_object)
    student["evaluations"] = evaluations
    return student

def get_heading_row(students):
    # (Nom, 2)

    heading = [("Nom", 0)]
    student = students[0]

    # get student evals
    for eval in student["evaluations"]:
        heading.append(
            (eval["criteria"], len(eval["values"])-1)
        )
    return heading

def get_row_for_student(student):

    # [nom, 5, 8, 3, 1]
    row = [student["name"]]
    for e in student["evaluations"]:
        row.extend(e["values"])    
    return row


# app views or endpoints

@app.route("/api/export/", methods=["POST", "GET"])
def export_data():
    students = db.all()
    if len(students) < 1:
        return redirect(url_for("home"))
    
    # workbook (excel file)
    wb = Workbook()
    ws = wb.active # excel sheet
    
    # heading
    heads = get_heading_row(students) # [(name, 0), (criteria1, 4), (criteria2, 4), ...]
    cell_index = 1
    for h in heads:
        ws.cell(row=1, column=cell_index).value = h[0]
        ws.merge_cells(start_row=1, start_column=cell_index, end_row=1, end_column=cell_index+h[1])
        cell_index = cell_index + h[1] + 1
    
    # place rows
    for i in range(1, len(students)+1):
        row = get_row_for_student(students[i-1])

        for j in range(len(row)):  # [nom, 5, 8, 3, 1]
            ws.cell(row=i+1, column=j+1).value = row[j]

    filename = Path(__file__).parent / "evaluations.xlsx"
    wb.save(filename=filename)
    return send_file(filename, as_attachment=True) # return excel file 

@app.route("/api/add/student/", methods=["GET", "POST"])
def add_student():
    if not request.method == "POST": # request method
        return redirect(url_for("home"))
    
    name = request.form.get("name")
    if name is None:
        name = (request.json).get("name")

    student = {
        "id": int(time.time()),
        "name": name,
        "evaluations": []
    }
    student["id"] = db.upsert(table.Document(student, doc_id=student["id"]))

    rq_status = status["success"]
    rq_status["object"] = student
    return rq_status

@app.route("/api/add/criteria/", methods=["GET", "POST"])
def add_criteria():
    if not request.method == "POST": # only POST request authorized
        return redirect(url_for("home"))
    
    # for forms
    name = request.form.get("name")
    datatype = request.form.get("datatype")
    
    if name is None:
        # for json data
        name = (request.json).get("name")
        datatype = (request.json).get("datatype")

    # criteria object
    criteria = {
        "id": int(time.time()),
        "name": name,
        "data-type": datatype
    }
    criteria["id"] = db_criteria.upsert(table.Document(criteria, doc_id=criteria["id"]))

    rq_status = status["success"]
    rq_status["object"] = criteria
    return rq_status # return request status

@app.route("/api/clear/", methods=["GET", "POST"])
def clear_db():
    db.truncate()
    db_criteria.truncate()
    return redirect(url_for("home"))

@app.route("/api/evaluate/", methods=["POST"])
def evaluate_student():
    """valid request body (json) 

        {
            'student_id': int
            'criteria_id': int
            'value': int
        }
    """

    # get data and deserialize them
    data = json.loads(request.json)

    student = db.get(doc_id=data.get("student_id")) # get student to evaluate
    if student: # student exist verification
        criteria = db_criteria.get(doc_id=data.get("criteria_id"))
        value = data.get("value")
        
        student = evaluate(student, criteria, value) # add evaluation value
        db.update(student, doc_ids=[data.get("student_id")]) # update student en database

    rq_status = status["success"]
    rq_status["object"] = student
    return rq_status

@app.route("/api/get/students/", methods=["GET", "POST"])
def get_student():
    return db.all()

@app.route("/api/get/criteria/", methods=["GET", "POST"])
def get_criteria():
    return db_criteria.all()

@app.route("/evaluations-list/")
def evaluation_list():
    data = db.all() # get all data

    table_head = None
    if len(data) > 1:
        table_head = get_heading_row(students=data)

    table_rows = []
    for student in data:
        row = get_row_for_student(student)
        table_rows.append(row)

    return render_template("evals_list.html", table_head=table_head, table_rows=table_rows)

@app.route("/")
def home():
    return render_template("home.html")


if __name__ == "__main__":
    app.run(debug=True)

